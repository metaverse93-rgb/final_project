"""
평가 결과 → Excel 리포트 생성
results_300.csv → eval_report_YYYYMMDD.xlsx (3시트)

실행:
    python eval/report.py
    python eval/report.py --before before.csv --after after.csv  # 파인튜닝 전/후 비교
"""

import os
import sys
import csv
import argparse
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")


def load_results(path: str) -> list[dict]:
    with open(path, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def generate_report(
    before_path: str,
    after_path: str = None,
    out_dir: str = None,
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("pip install openpyxl")

    before = load_results(before_path)
    after  = load_results(after_path) if after_path else None

    wb = openpyxl.Workbook()

    # ── 스타일 정의 ──
    H_FILL  = PatternFill("solid", start_color="2D5986", end_color="2D5986")
    H_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    D_FONT  = Font(name="Arial", size=9)
    CENTER  = Alignment(horizontal="center", vertical="center")
    WRAP    = Alignment(wrap_text=True, vertical="top")

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = H_FILL
            cell.font = H_FONT
            cell.alignment = CENTER
        ws.row_dimensions[1].height = 20

    # ── Sheet 1: 번역 평가 ──
    ws1 = wb.active
    ws1.title = "번역 평가"
    headers1 = ["NO", "카테고리", "URL", "원문(EN)", "GT(KO)", "번역 출력", "BLEU", "COMET", "TPR", "누락 용어"]
    style_header(ws1, headers1)
    for ri, row in enumerate(before, 2):
        ws1.cell(ri, 1, row["id"]).font = D_FONT
        ws1.cell(ri, 2, row["category"]).font = D_FONT
        ws1.cell(ri, 3, row["url"]).font = D_FONT
        ws1.cell(ri, 4, row["en_text"]).alignment = WRAP
        ws1.cell(ri, 5, row["ko_gt"]).alignment = WRAP
        ws1.cell(ri, 6, row["translation"]).alignment = WRAP
        ws1.cell(ri, 7, float(row["bleu"] or 0)).font = D_FONT
        ws1.cell(ri, 8, float(row.get("comet") or 0)).font = D_FONT
        ws1.cell(ri, 9, float(row["tpr"] or 0)).font = D_FONT
        ws1.cell(ri, 10, row["tpr_missing"]).font = D_FONT

    col_widths1 = [6, 14, 40, 50, 50, 50, 8, 10, 8, 30]
    for i, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: 요약 평가 (격식체) ──
    ws2 = wb.create_sheet("요약 평가 (격식체)")
    headers2 = ["NO", "카테고리", "원문(EN)", "격식체 요약",
                "충실성", "유창성", "간결성", "관련성",
                "G-Eval 단순평균", "G-Eval 가중평균"]
    style_header(ws2, headers2)
    for ri, row in enumerate(before, 2):
        ws2.cell(ri, 1, row["id"]).font = D_FONT
        ws2.cell(ri, 2, row["category"]).font = D_FONT
        ws2.cell(ri, 3, row["en_text"]).alignment = WRAP
        ws2.cell(ri, 4, row["summary_formal"]).alignment = WRAP
        ws2.cell(ri, 5, int(float(row.get("geval_consistency") or 0))).font = D_FONT
        ws2.cell(ri, 6, int(float(row.get("geval_fluency") or 0))).font = D_FONT
        ws2.cell(ri, 7, int(float(row.get("geval_coherence") or 0))).font = D_FONT
        ws2.cell(ri, 8, int(float(row.get("geval_relevance") or 0))).font = D_FONT
        ws2.cell(ri, 9,  float(row.get("g_eval_score") or 0)).font = D_FONT
        ws2.cell(ri, 10, float(row.get("g_eval_weighted") or 0)).font = D_FONT

    col_widths2 = [6, 14, 50, 50, 10, 10, 10, 10, 14, 14]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: 종합 요약 (파인튜닝 전/후 비교) ──
    ws3 = wb.create_sheet("종합 요약")

    def safe_mean(rows, key):
        vals = [float(r[key]) for r in rows if r.get(key) and float(r[key]) > 0]
        return round(sum(vals) / len(vals), 2) if vals else 0.0

    metrics = [
        ("BLEU",              "bleu",                "≥ 17.0"),
        ("COMET",             "comet",               "기준값 측정 후"),
        ("TPR",               "tpr",                 "≥ 0.95"),
        ("G-Eval 일치성",      "geval_consistency",   "≥ 4.0"),
        ("G-Eval 유창성",      "geval_fluency",       "≥ 4.0"),
        ("G-Eval 일관성",      "geval_coherence",     "≥ 4.0"),
        ("G-Eval 관련성",      "geval_relevance",     "≥ 4.0"),
        ("G-Eval 단순평균",    "g_eval_score",        "≥ 4.0"),
        ("G-Eval 가중평균",    "g_eval_weighted",     "≥ 4.0"),
    ]

    headers3 = ["지표", "파인튜닝 전", "파인튜닝 후", "목표값", "달성 여부"]
    style_header(ws3, headers3)

    for ri, (label, key, target) in enumerate(metrics, 2):
        before_val = safe_mean(before, key)
        after_val  = safe_mean(after, key) if after else "-"

        if after and isinstance(after_val, float) and target.startswith("≥"):
            goal = float(target.replace("≥", "").strip())
            achieved = "✅" if after_val >= goal else "❌"
        else:
            achieved = "-"

        ws3.cell(ri, 1, label).font = D_FONT
        ws3.cell(ri, 2, before_val).font = D_FONT
        ws3.cell(ri, 3, after_val).font = D_FONT
        ws3.cell(ri, 4, target).font = D_FONT
        ws3.cell(ri, 5, achieved).font = D_FONT

    col_widths3 = [18, 14, 14, 18, 12]
    for i, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── 저장 ──
    if out_dir is None:
        out_dir = os.path.dirname(before_path)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = os.path.join(out_dir, f"eval_report_{ts}.xlsx")
    wb.save(out_path)
    print(f"리포트 저장 완료: {out_path}")
    return out_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--before", default=os.path.join(os.path.dirname(__file__), "data", "results_300.csv"))
    parser.add_argument("--after",  default=None, help="파인튜닝 후 results.csv 경로")
    args = parser.parse_args()

    generate_report(before_path=args.before, after_path=args.after)
