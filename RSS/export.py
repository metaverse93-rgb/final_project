"""
三鮮 (삼선) - DB 엑셀 내보내기
사용법:
    python export.py                  # samsun.db → samsun_articles.xlsx
    python export.py --db my.db       # DB 경로 지정
    python export.py --out result.xlsx # 출력 파일명 지정

설치:
    pip install openpyxl
"""

import sqlite3
import argparse
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def export_to_excel(db_path: str = "samsun.db", out_path: str = None):
    if out_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        out_path = f"samsun_articles_{ts}.xlsx"

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    rows = conn.execute("""
        SELECT id, source, source_type, category, country,
               title, url, content, credibility_score, published_at, created_at
        FROM articles
        ORDER BY source_type, source, published_at DESC
    """).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "articles"

    headers = [
        "id", "source", "source_type", "category", "country",
        "title", "url", "content", "credibility_score", "published_at", "created_at"
    ]

    # 헤더 스타일
    header_fill = PatternFill("solid", start_color="2D5986", end_color="2D5986")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 데이터 행
    row_font = Font(name="Arial", size=9)
    community_fill = PatternFill("solid", start_color="EBF5EB", end_color="EBF5EB")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
            cell.font = row_font
            cell.alignment = Alignment(
                wrap_text=(key == "content"),
                vertical="top"
            )
        # 커뮤니티 행 배경색 구분
        if row["source_type"] == "community":
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = community_fill

    # 컬럼 너비
    col_widths = [6, 22, 12, 14, 10, 50, 40, 60, 16, 20, 20]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(rows) + 1}"

    wb.save(out_path)
    print(f"✅ 내보내기 완료: {out_path} ({len(rows)}건)")
    return out_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="三鮮 DB → Excel 내보내기")
    parser.add_argument("--db",  default="samsun.db",   help="DB 파일 경로")
    parser.add_argument("--out", default=None,           help="출력 파일명 (기본: samsun_articles_YYYYMMDD_HHMM.xlsx)")
    args = parser.parse_args()

    export_to_excel(args.db, args.out)
